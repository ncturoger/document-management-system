from factory import create_app
from database import db
from db_model import Post
import click
import openpyxl
from bs4 import BeautifulSoup
import requests



app = create_app()


@app.cli.command()
def test():
    click.echo("test")


@app.cli.command()
def test_insert():
    data = Post('test', '123')
    db.session.add(data)
    db.session.commit()

@app.cli.command()
@click.option('--excel_path', prompt="Enter the excel path")
def test_insert_doc(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    post_list = []

    sheetnames = wb.sheetnames

    for sheet in sheetnames:
        ws = wb[sheet]
    
        for row in list(ws.rows)[1:]:
            if len(row) < 2 or not row[1].hyperlink:
                break

            post_list.append({'title': row[1].value, 'link': row[1].hyperlink.target})
    
    # print(post_list)
    for post in post_list:
        res = requests.get(post['link'])
        # print(res.text)
        soup = BeautifulSoup(res.text, 'html.parser')
        content = soup.find('div', {'class': 'post_text'})
        if content:
            post_content = content.text
            post_record = Post(post['title'], post_content)
            db.session.add(post_record)
            db.session.commit()
        # soup_body = soup.find('body')
        # for div_tag in soup_body.find_all('div'):
        #     print(div_tag.text)


